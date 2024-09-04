from django.shortcuts import render,get_object_or_404,redirect
from django.contrib.auth.decorators import login_required
from accounts.forms import UserChangeForm, AddressForm
from accounts.models import Address
from facades.views import InformationsForTemplate
from .forms import *
from .models import ExcelStuff
from django.utils.text import slugify
from django.db import IntegrityError
from django.db.models import Q
from orders.models import Order
from config.settings import SMS_PASSWORD
#----------------------------------------------------------------------------------------------
@login_required
def dashboard(request,address_id = None):
    profileForm = UserChangeForm(instance=request.user)
    addressForm = AddressForm()
    Info = InformationsForTemplate(request)

    if(address_id):
        address = get_object_or_404(Address,id=address_id)
        EditAddressForm = AddressForm(instance=address)
        Info["EditAddressForm"] = EditAddressForm
        Info["EditAddressId"] = address_id

    
    if request.user.is_admin:
        Info["orders_pending_confirmation"] = Order.objects.filter(processed=False)
        Info["orders_not_confirm"] = Order.objects.filter(not_confirm=True)
        Info["orders_pending_packing"] = Order.objects.filter(Q(packing=False) & Q(processed=True))
        Info["orders_deliveried"] = Order.objects.filter(Q(deliveried=False) & Q(packing=True))
        Info["orders_final"] = Order.objects.filter(Q(deliveried=True) & Q(packing=True))
        
        

    
    Info["profileForm"] = profileForm
    Info["addressForm"] = addressForm
    Info["excelForm"] = ExcelForm() 
    Info["last_id"] = Product.objects.all().order_by('-id')[0].id

    return render(request,'administratorship/dashboard.html',Info)

#----------------------------------------------------------------------------------------------
def check_product_list(request,id):
    Info = InformationsForTemplate(request)
    Info["order"] = Order.objects.get(id=id)

    
    return render(request,'administratorship/ProductList.html',Info)
#----------------------------------------------------------------------------------------------
def order_not_confirm(request,id):
    if request.user.is_admin:
        order = Order.objects.get(id=id)
        order.not_confirm = True
        sms = Client(SMS_PASSWORD)

        # send sms to Customer
        user = order.user

        pattern_values = {
            "username": str(user.full_name),
        }

        message_id = sms.send_pattern(
            "31vo2au2smdjjqf",    # pattern code
            "+983000505",      # originator
            f"98{user.phoneNumber[1:]}",  # recipient
            pattern_values,  # pattern values
        )
        print(f"message sended to customer id: {message_id}")

        order.save()
    return redirect("administratorship:dashboard")
#----------------------------------------------------------------------------------------------
def order_confirmation(request,id):
    if request.user.is_admin:
        order = Order.objects.get(id=id)
        order.processed = True

        # send sms to Customer
        user = order.user
        sms = Client(SMS_PASSWORD)

        pattern_values = {
            "username": str(user.full_name),
            "code": str(order.get_order_number())
        }

        message_id = sms.send_pattern(
            "j6y8q4md5ft1uhj",    # pattern code
            "+983000505",      # originator
            f"98{user.phoneNumber[1:]}",  # recipient
            pattern_values,  # pattern values
        )
        print(f"message sended to customer id: {message_id}")

        order.save()
    return redirect("administratorship:dashboard")
#----------------------------------------------------------------------------------------------
def order_packing_shipping_confirm(request, id):
    if request.user.is_admin:
        order = Order.objects.get(id=id)
        
        order.interception_code = request.POST.get('check_code')
        order.packing = True
        order.shipped = True

        order.save()

        sms = Client(SMS_PASSWORD)
        # send sms to Customer
        user = order.user

        pattern_values = {
            "username": str(user.full_name),
            "code": str(order.get_order_number()),
            "postcode": str(order.interception_code)
        }

        message_id = sms.send_pattern(
            "kcsdvcjzv988oa4",    # pattern code
            "+983000505",      # originator
            f"98{user.phoneNumber[1:]}",  # recipient
            pattern_values,  # pattern values
        )
        print(f"message sended to customer id: {message_id}")
        
        
    return redirect("administratorship:dashboard")
     
#----------------------------------------------------------------------------------------------
def order_delivered_confirm(request,id):
    if request.user.is_admin:
        order = Order.objects.get(id=id)
        order.deliveried = True
        # send sms

        order.save()
    return redirect("administratorship:dashboard")        
#----------------------------------------------------------------------------------------------
import openpyxl
from stuff.models import Product,Category,Brand,Color
# @csrf_exemp
def add_group_of_product(request):
    if request.method == 'POST':
        form = ExcelForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            
            # Opening excel File
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                print(row[1])
                id = row[0]
                name = row[1]
                en = row[2]
                category_names = row[3].split(',') 
                brand_name = row[4]
                price = row[5]
                warehouse = row[6]
                product_barcode = row[7]
                short_description = row[8]
                description = row[9]
                more_info = row[10]
                rating =  row[11]
                slug = slugify(row[12])
                alt = row[13]
                available = True if(warehouse > 0) else False
                colors = row[14].split(',')  
                dirham_price = row[15]
                transit_price = row[16]
                image = "row[11]"


                print("--------------------")

                
                # Check if product already exists
                try:
                    product = Product.objects.get(slug=slug,id=id)
                    # Update existing product
                    product.name = name
                    product.En = en
                    product.price = price
                    product.warehouse = warehouse
                    product.short_description = short_description
                    product.description = description
                    product.more_info = more_info
                    product.rating = rating
                    product.alt = alt
                    product.available = available
                    product.product_barcode = product_barcode
                    product.dirham_price = dirham_price
                    product.transit_price = transit_price
                    product.save()

                except Product.DoesNotExist:

                    try:
                        product = Product.objects.create(
                            name=name,
                            slug=slug,
                            price=price,
                            En=en,
                            image=image,
                            alt=alt,
                            available=available,
                            rating=rating,
                            warehouse=warehouse,
                            short_description=short_description,
                            description=description,
                            more_info=more_info,
                            product_barcode=product_barcode,
                            dirham_price=dirham_price,
                            transit_price=transit_price
                        )
                    except IntegrityError:
                        # Handle the unique constraint violation
                        print("Unique constraint violation: Product with the same ID or slug already exists.")
                
                
                # Add or Create Category
                for category_name in category_names:
                    category, _ = Category.objects.get_or_create(name=category_name)       
                    product.category.add(category)
                
                # Add or Create Brand
                brand, _ = Brand.objects.get_or_create(name=brand_name)
                product.brand = brand
                
                # Add or Create Color
                for color_name in colors:
                    color, _ = Color.objects.get_or_create(name=color_name)
                    product.colors.add(color)
                
                product.save()
            
            return redirect("facades:home")

    return dashboard(request)
#----------------------------------------------------------------------------------------------
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
def export_products_to_excel(reqeust):
    products = Product.objects.all().order_by('id')
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    
    sheet[f"A1"] = "شناسه"
    sheet[f"B1"] = "نام"
    sheet[f"C1"] = "نام انگلیسی"
    sheet[f"D1"] = "نام دسته بندی"
    sheet[f"E1"] = "نام برند"
    sheet[f"f1"] = "قیمت قبل از تخفیف (تومان)"
    sheet[f"G1"] = "موجودی"
    sheet[f"H1"] = "بارکد"

    sheet[f"I1"] = "توضیحات کوتاه"
    sheet[f"J1"] = "توضیحات"
    sheet[f"K1"] = "توضیحات بیشتر"
    sheet[f"L1"] = "امتیاز"
    sheet[f"M1"] = "اسلاگ"
    sheet[f"N1"] = "متن عکس"
    sheet[f"O1"] = "رنگ ها"
    sheet[f"P1"] = "قیمت درهمی"
    sheet[f"Q1"] = "قیمت ارسال"


    for row_num, product in enumerate(products, start=2):
        sheet[f"A{row_num}"] = product.id
        sheet[f"B{row_num}"] = product.name
        sheet[f"C{row_num}"] = product.En
        sheet[f"D{row_num}"] = ",".join([str(category) for category in product.category.all()])
        sheet[f"E{row_num}"] = product.brand.name
        sheet[f"F{row_num}"] = product.price
        sheet[f"G{row_num}"] = product.warehouse
        sheet[f"H{row_num}"] = product.product_barcode

        sheet[f"I{row_num}"] = product.short_description
        sheet[f"J{row_num}"] = product.description
        sheet[f"K{row_num}"] = product.more_info
        sheet[f"L{row_num}"] = product.rating
        sheet[f"M{row_num}"] = product.slug
        sheet[f"N{row_num}"] = product.alt
        sheet[f"P{row_num}"] = product.dirham_price
        sheet[f"Q{row_num}"] = product.transit_price

        colors = product.colors.all()
        if(len(colors) > 0):
            colors_text = f"{colors[0].name}"
            for color in colors[1:]:
                colors_text += "," + color.name     

        sheet[f"O{row_num}"] = colors_text


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="products.xlsx"'

    workbook.save(response)

    return response

