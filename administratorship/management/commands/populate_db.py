from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from django.utils.text import slugify
from stuff.models import Category
import os

class Command(BaseCommand):
    help = 'Populate the database with initial data'

    def handle(self, *args, **options):

        #path
        current_directory = os.getcwd()
        excel_file_path = os.path.join(current_directory, 'excels/categories.xlsx')
        
        #Read from excel
        workbook = load_workbook(filename=excel_file_path)
        sheet = workbook.active

        #fill database
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = row[0]
            is_sub = row[1]
            subcategory_name = row[2] if len(row) > 2 else None

            data = {
                'name': name,
                'slug': slugify(name),
                'is_sub': is_sub,
                'sub_category': Category.objects.get(name=subcategory_name) if subcategory_name else None
            }
            Category.objects.create(**data)

        self.stdout.write(self.style.SUCCESS('Database populated successfully.'))